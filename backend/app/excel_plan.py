from __future__ import annotations

import base64
import binascii
import hashlib
import json
import re
import sqlite3
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .database import row_to_dict


MAX_WORKBOOK_BYTES = 10 * 1024 * 1024
TRUE_MARKS = {"1", "true", "yes", "y", "done", "complete", "completed", "✓", "✔", "☑"}
FALSE_MARKS = {"0", "false", "no", "n", "missed", "skipped", "✗", "✘", "☒"}


def decode_workbook_payload(encoded: str) -> bytes:
    value = encoded.split(",", 1)[1] if encoded.startswith("data:") and "," in encoded else encoded
    try:
        data = base64.b64decode(value, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError("Workbook must be valid base64 data.") from exc
    if not data:
        raise ValueError("Workbook is empty.")
    if len(data) > MAX_WORKBOOK_BYTES:
        raise ValueError("Workbook is larger than the 10 MB import limit.")
    return data


def parse_workbook_plan(data: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(data), data_only=False, read_only=True)
    except Exception as exc:
        raise ValueError("The uploaded file is not a readable .xlsx workbook.") from exc

    sheet = next((item for item in workbook.worksheets if _find_phase_header(item)), None)
    if sheet is None:
        raise ValueError("Could not find the phase table in this workbook.")

    phase_header_row = _find_phase_header(sheet)
    phases = _parse_phases(sheet, phase_header_row)
    if not phases:
        raise ValueError("No Phase 1-5 rows were found in the workbook.")

    title_lines = [str(sheet.cell(row=row, column=1).value or "") for row in range(1, phase_header_row)]
    profile = _parse_profile(" ".join(title_lines), phases)
    tracking_weeks = _count_tracking_weeks(sheet)
    rules = _parse_rules(sheet)
    history_rows = _parse_history_rows(sheet)

    return {
        "sheet_name": sheet.title,
        "source_sha256": hashlib.sha256(data).hexdigest(),
        "profile": profile,
        "phases": phases,
        "tracking_weeks": tracking_weeks,
        "rules": rules,
        "history_rows": history_rows,
    }


def preview_workbook_plan(parsed: dict[str, Any], filename: str, start_date: date) -> dict[str, Any]:
    weeks = parsed["tracking_weeks"]
    cycle_end = start_date + timedelta(days=max(weeks * 7 - 1, 0))
    return {
        "filename": filename,
        "sheet_name": parsed["sheet_name"],
        "source_sha256": parsed["source_sha256"],
        "plan_start_date": start_date.isoformat(),
        "cycle_end_date": cycle_end.isoformat(),
        "tracking_weeks": weeks,
        "profile": parsed["profile"],
        "phases": parsed["phases"],
        "rules": parsed["rules"],
        "history_rows_found": len(parsed["history_rows"]),
        "warnings": [
            "Phase changes remain weight-trend based; the 8-week calendar is a tracking cycle.",
            "Blank cells and empty checkboxes remain unknown and are not imported as zero or false.",
        ],
    }


def apply_workbook_plan(
    conn: sqlite3.Connection,
    parsed: dict[str, Any],
    filename: str,
    start_date: date,
) -> dict[str, Any]:
    existing = conn.execute(
        "SELECT * FROM plan_imports WHERE source_sha256 = ? AND plan_start_date = ?",
        (parsed["source_sha256"], start_date.isoformat()),
    ).fetchone()
    if existing is not None:
        return {
            "applied": False,
            "already_applied": True,
            "import_id": existing["id"],
            "timeline": get_goal_timeline(conn, start_date),
        }

    summary = preview_workbook_plan(parsed, filename, start_date)
    cursor = conn.execute(
        """
        INSERT INTO plan_imports (
            source_name, source_sha256, sheet_name, plan_start_date,
            tracking_weeks, summary_json
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            filename,
            parsed["source_sha256"],
            parsed["sheet_name"],
            start_date.isoformat(),
            parsed["tracking_weeks"],
            json.dumps(summary),
        ),
    )
    import_id = int(cursor.lastrowid)

    phase_ids: dict[int, int] = {}
    for phase in parsed["phases"]:
        existing_phase = conn.execute("SELECT * FROM phases WHERE name = ?", (phase["name"],)).fetchone()
        if existing_phase is not None:
            revision_count = conn.execute(
                "SELECT COUNT(*) AS count FROM goal_revisions WHERE phase_id = ?",
                (existing_phase["id"],),
            ).fetchone()["count"]
            if revision_count == 0:
                conn.execute(
                    """
                    INSERT INTO goal_revisions (
                        phase_id, effective_date, start_weight_kg, end_weight_kg,
                        calorie_target, calorie_min, calorie_max, protein_min_g,
                        protein_max_g, steps_target, weekday_run_km, sunday_run_km,
                        workout_days_per_week, primary_goal, source_import_id, reason
                    ) VALUES (?, '0001-01-01', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?)
                    """,
                    (
                        existing_phase["id"],
                        existing_phase["start_weight_kg"],
                        existing_phase["end_weight_kg"],
                        existing_phase["calorie_target"],
                        existing_phase["calorie_min"],
                        existing_phase["calorie_max"],
                        existing_phase["protein_min_g"],
                        existing_phase["protein_max_g"],
                        existing_phase["steps_target"],
                        existing_phase["weekday_run_km"],
                        existing_phase["sunday_run_km"],
                        existing_phase["workout_days_per_week"],
                        existing_phase["primary_goal"],
                        "Baseline before first workbook import",
                    ),
                )
        conn.execute(
            """
            INSERT INTO phases (
                name, order_index, start_weight_kg, end_weight_kg,
                calorie_target, calorie_min, calorie_max, protein_min_g,
                protein_max_g, steps_target, weekday_run_km, sunday_run_km,
                workout_days_per_week, primary_goal
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                order_index = excluded.order_index,
                start_weight_kg = excluded.start_weight_kg,
                end_weight_kg = excluded.end_weight_kg,
                calorie_target = excluded.calorie_target,
                calorie_min = excluded.calorie_min,
                calorie_max = excluded.calorie_max,
                protein_min_g = excluded.protein_min_g,
                protein_max_g = excluded.protein_max_g,
                steps_target = excluded.steps_target,
                weekday_run_km = excluded.weekday_run_km,
                sunday_run_km = excluded.sunday_run_km,
                workout_days_per_week = excluded.workout_days_per_week,
                primary_goal = excluded.primary_goal
            """,
            (
                phase["name"],
                phase["order_index"],
                phase["start_weight_kg"],
                phase["end_weight_kg"],
                phase["calorie_target"],
                phase["calorie_min"],
                phase["calorie_max"],
                phase["protein_min_g"],
                phase["protein_max_g"],
                phase["steps_target"],
                phase["weekday_run_km"],
                phase["sunday_run_km"],
                phase["workout_days_per_week"],
                phase["primary_goal"],
            ),
        )
        row = conn.execute("SELECT id FROM phases WHERE name = ?", (phase["name"],)).fetchone()
        phase_id = int(row["id"])
        phase_ids[phase["order_index"]] = phase_id
        conn.execute(
            """
            INSERT INTO goal_revisions (
                phase_id, effective_date, start_weight_kg, end_weight_kg,
                calorie_target, calorie_min, calorie_max, protein_min_g,
                protein_max_g, steps_target, weekday_run_km, sunday_run_km,
                workout_days_per_week, primary_goal, source_import_id, reason
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                phase_id,
                start_date.isoformat(),
                phase["start_weight_kg"],
                phase["end_weight_kg"],
                phase["calorie_target"],
                phase["calorie_min"],
                phase["calorie_max"],
                phase["protein_min_g"],
                phase["protein_max_g"],
                phase["steps_target"],
                phase["weekday_run_km"],
                phase["sunday_run_km"],
                phase["workout_days_per_week"],
                phase["primary_goal"],
                import_id,
                f"Imported from {filename}",
            ),
        )

    profile = parsed["profile"]
    conn.execute(
        """
        UPDATE user_profile SET
            start_weight_kg = ?, goal_weight_min_kg = ?, goal_weight_max_kg = ?,
            current_phase_id = COALESCE(current_phase_id, ?),
            updated_at = CURRENT_TIMESTAMP
        WHERE id = 1
        """,
        (
            profile["start_weight_kg"],
            profile["goal_weight_min_kg"],
            profile["goal_weight_max_kg"],
            phase_ids[min(phase_ids)],
        ),
    )
    settings = {
        "plan_start_date": start_date.isoformat(),
        "tracking_cycle_weeks": parsed["tracking_weeks"],
        "plan_source_name": filename,
        "plan_source_sha256": parsed["source_sha256"],
        "plan_rules": parsed["rules"],
    }
    for key, value in settings.items():
        conn.execute(
            """
            INSERT INTO settings (key, value_json) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value_json = excluded.value_json
            """,
            (key, json.dumps(value)),
        )

    imported_logs = _apply_history_rows(conn, parsed["history_rows"])
    conn.commit()
    return {
        "applied": True,
        "already_applied": False,
        "import_id": import_id,
        "daily_logs_imported": imported_logs,
        "timeline": get_goal_timeline(conn, start_date),
    }


def resolve_phase_targets(conn: sqlite3.Connection, phase: sqlite3.Row, on_date: date) -> dict[str, Any]:
    resolved = row_to_dict(phase) or {}
    revision = conn.execute(
        """
        SELECT * FROM goal_revisions
        WHERE phase_id = ? AND effective_date <= ?
        ORDER BY effective_date DESC, id DESC
        LIMIT 1
        """,
        (phase["id"], on_date.isoformat()),
    ).fetchone()
    if revision is None:
        return resolved
    for key in (
        "start_weight_kg",
        "end_weight_kg",
        "calorie_target",
        "calorie_min",
        "calorie_max",
        "protein_min_g",
        "protein_max_g",
        "steps_target",
        "weekday_run_km",
        "sunday_run_km",
        "workout_days_per_week",
        "primary_goal",
    ):
        resolved[key] = revision[key]
    resolved["goal_revision_id"] = revision["id"]
    resolved["goals_effective_date"] = revision["effective_date"]
    return resolved


def get_goal_timeline(conn: sqlite3.Connection, on_date: date) -> dict[str, Any]:
    profile = conn.execute("SELECT * FROM user_profile WHERE id = 1").fetchone()
    phase = conn.execute(
        """
        SELECT phases.* FROM user_profile
        JOIN phases ON phases.id = user_profile.current_phase_id
        WHERE user_profile.id = 1
        """
    ).fetchone()
    if profile is None or phase is None:
        raise ValueError("Profile or current phase is not configured.")

    latest_import = conn.execute("SELECT * FROM plan_imports ORDER BY id DESC LIMIT 1").fetchone()
    cycle = None
    if latest_import is not None:
        start = date.fromisoformat(latest_import["plan_start_date"])
        weeks = int(latest_import["tracking_weeks"])
        end = start + timedelta(days=weeks * 7 - 1)
        elapsed = (on_date - start).days
        cycle = {
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "tracking_weeks": weeks,
            "current_week": elapsed // 7 + 1 if 0 <= elapsed < weeks * 7 else None,
            "current_day": elapsed + 1 if 0 <= elapsed < weeks * 7 else None,
            "status": "upcoming" if elapsed < 0 else "complete" if elapsed >= weeks * 7 else "active",
        }

    phases = [row_to_dict(row) for row in conn.execute("SELECT * FROM phases ORDER BY order_index").fetchall()]
    return {
        "date": on_date.isoformat(),
        "source": {
            "filename": latest_import["source_name"],
            "sha256": latest_import["source_sha256"],
            "imported_at": latest_import["imported_at"],
        } if latest_import is not None else None,
        "cycle": cycle,
        "profile": row_to_dict(profile),
        "current_phase": resolve_phase_targets(conn, phase, on_date),
        "phases": phases,
        "phase_transition_policy": "weight_trend_with_manual_review",
    }


def _find_phase_header(sheet: Any) -> int | None:
    for row in range(1, min(sheet.max_row, 30) + 1):
        values = [str(sheet.cell(row=row, column=column).value or "").strip().lower() for column in range(1, 8)]
        if values[:3] == ["phase", "weight range", "calories*"] or (
            values[0] == "phase" and "weight" in values[1] and "calorie" in values[2]
        ):
            return row
    return None


def _parse_phases(sheet: Any, header_row: int) -> list[dict[str, Any]]:
    phases: list[dict[str, Any]] = []
    previous_run = (2.0, 5.0)
    for row in range(header_row + 1, min(sheet.max_row, header_row + 15) + 1):
        name = str(sheet.cell(row=row, column=1).value or "").strip()
        match = re.fullmatch(r"Phase\s+(\d+)", name, re.IGNORECASE)
        if not match:
            continue
        order = int(match.group(1))
        weight_text = str(sheet.cell(row=row, column=2).value or "")
        calorie_text = str(sheet.cell(row=row, column=3).value or "")
        protein_text = str(sheet.cell(row=row, column=4).value or "")
        steps_text = str(sheet.cell(row=row, column=5).value or "")
        run_text = str(sheet.cell(row=row, column=6).value or "")
        goal = str(sheet.cell(row=row, column=7).value or "").strip()

        start_weight, end_weight = _weight_range(weight_text, maintenance=order == 5)
        calorie_min, calorie_max, calorie_target = _calorie_range(calorie_text)
        protein_numbers = _numbers(protein_text)
        protein_min = int(min(protein_numbers)) if protein_numbers else 0
        protein_max = int(max(protein_numbers)) if protein_numbers else protein_min
        steps_target = _steps_target(steps_text)
        parsed_run = _run_targets(run_text)
        if parsed_run is not None:
            previous_run = parsed_run

        phases.append(
            {
                "name": f"Phase {order}",
                "order_index": order,
                "start_weight_kg": start_weight,
                "end_weight_kg": end_weight,
                "calorie_target": calorie_target,
                "calorie_min": calorie_min,
                "calorie_max": calorie_max,
                "protein_min_g": protein_min,
                "protein_max_g": protein_max,
                "steps_target": steps_target,
                "weekday_run_km": previous_run[0],
                "sunday_run_km": previous_run[1],
                "workout_days_per_week": 6,
                "primary_goal": goal,
                "source_labels": {
                    "weight": weight_text,
                    "calories": calorie_text,
                    "protein": protein_text,
                    "steps": steps_text,
                    "running": run_text,
                },
            }
        )
    return sorted(phases, key=lambda item: item["order_index"])


def _parse_profile(text: str, phases: list[dict[str, Any]]) -> dict[str, float]:
    match = re.search(r"Goal:\s*([\d.]+)\s*kg\s*[→-]\s*([\d.]+)\s*[–-]\s*([\d.]+)", text)
    if match:
        return {
            "start_weight_kg": float(match.group(1)),
            "goal_weight_min_kg": float(match.group(2)),
            "goal_weight_max_kg": float(match.group(3)),
        }
    return {
        "start_weight_kg": phases[0]["start_weight_kg"],
        "goal_weight_min_kg": phases[-1]["end_weight_kg"],
        "goal_weight_max_kg": phases[-1]["start_weight_kg"],
    }


def _count_tracking_weeks(sheet: Any) -> int:
    weeks = []
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row=row, column=1).value or "").strip()
        match = re.fullmatch(r"Week\s+(\d+)", value, re.IGNORECASE)
        if match:
            weeks.append(int(match.group(1)))
    return max(weeks, default=0)


def _parse_rules(sheet: Any) -> list[str]:
    rules: list[str] = []
    in_rules = False
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row=row, column=1).value or "").strip()
        if value.upper() == "PHASE / ADJUSTMENT RULES":
            in_rules = True
            continue
        if in_rules and value:
            rules.append(value.lstrip("• ").strip())
    return rules


def _parse_history_rows(sheet: Any) -> list[dict[str, Any]]:
    header_row = None
    for row in range(1, sheet.max_row + 1):
        first = str(sheet.cell(row=row, column=1).value or "").strip().lower()
        second = str(sheet.cell(row=row, column=2).value or "").strip().lower()
        if first == "day" and second == "date":
            header_row = row
            break
    if header_row is None:
        return []

    results: list[dict[str, Any]] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        raw_date = sheet.cell(row=row, column=2).value
        local_date = _excel_date(raw_date, sheet.parent.epoch)
        if local_date is None:
            continue
        item: dict[str, Any] = {"local_date": local_date.isoformat()}
        weight = _optional_float(sheet.cell(row=row, column=3).value)
        if weight is not None:
            item["weight_kg"] = weight

        for column, field in (
            (7, "run_completed"),
            (8, "gym_completed"),
            (9, "breakfast_completed"),
            (10, "lunch_completed"),
            (11, "post_workout_completed"),
            (12, "dinner_completed"),
        ):
            checked = _checkbox(sheet.cell(row=row, column=column).value)
            if checked is not None:
                item[field] = checked

        sleep = _checkbox(sheet.cell(row=row, column=13).value)
        if sleep is True:
            item["sleep_hours"] = 8.0
        notes = sheet.cell(row=row, column=14).value
        if notes not in (None, ""):
            item["notes"] = str(notes)
        if len(item) > 1:
            results.append(item)
    return results


def _apply_history_rows(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    imported = 0
    allowed = {
        "weight_kg",
        "run_completed",
        "gym_completed",
        "breakfast_completed",
        "lunch_completed",
        "post_workout_completed",
        "dinner_completed",
        "sleep_hours",
        "notes",
    }
    bool_fields = {field for field in allowed if field.endswith("_completed")}
    for item in rows:
        fields = {key: value for key, value in item.items() if key in allowed}
        if not fields:
            continue
        conn.execute("INSERT OR IGNORE INTO daily_logs (local_date) VALUES (?)", (item["local_date"],))
        assignments = ", ".join(f"{key} = COALESCE(?, {key})" for key in fields)
        values = [1 if value is True else 0 if value is False and key in bool_fields else value for key, value in fields.items()]
        conn.execute(
            f"UPDATE daily_logs SET {assignments}, updated_at = CURRENT_TIMESTAMP WHERE local_date = ?",
            (*values, item["local_date"]),
        )
        imported += 1
    return imported


def _numbers(value: str) -> list[float]:
    return [float(item) for item in re.findall(r"\d+(?:\.\d+)?", value.replace(",", ""))]


def _weight_range(value: str, maintenance: bool = False) -> tuple[float, float]:
    numbers = _numbers(value)
    if len(numbers) < 2:
        raise ValueError(f"Could not understand weight range: {value}")
    if maintenance:
        return max(numbers[:2]), min(numbers[:2])
    # "76 → 72-73 kg" means the cut exits into the 72-73 maintenance
    # band. Use its upper edge as the transition threshold.
    return numbers[0], max(numbers[1:])


def _calorie_range(value: str) -> tuple[int | None, int | None, int | None]:
    if "maintenance" in value.lower():
        return None, None, None
    numbers = [int(item) for item in _numbers(value)]
    if not numbers:
        return None, None, None
    low, high = min(numbers), max(numbers)
    return low, high, int(round((low + high) / 2))


def _steps_target(value: str) -> int:
    numbers = _numbers(value)
    if not numbers:
        raise ValueError(f"Could not understand steps target: {value}")
    multiplier = 1000 if "k" in value.lower() else 1
    return int(numbers[0] * multiplier)


def _run_targets(value: str) -> tuple[float, float] | None:
    if "mon" not in value.lower() or "sun" not in value.lower():
        return None
    numbers = _numbers(value)
    if len(numbers) < 2:
        return None
    return numbers[0], numbers[1]


def _excel_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value, epoch).date()
    if isinstance(value, str) and value.strip():
        for parser in (date.fromisoformat, lambda item: datetime.strptime(item, "%d/%m/%Y").date()):
            try:
                return parser(value.strip())
            except ValueError:
                continue
    return None


def _checkbox(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in TRUE_MARKS:
        return True
    if normalized in FALSE_MARKS:
        return False
    return None


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
