from __future__ import annotations

import base64
from io import BytesIO
import json

from fastapi.testclient import TestClient
from openpyxl import Workbook


def make_client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("TRACKER_DB_PATH", str(tmp_path / "test.sqlite3"))
    monkeypatch.setenv("TRACKER_ALLOW_UNVERIFIED_GOOGLE", "1")
    monkeypatch.setenv("AUTH_RATE_LIMIT", "0")
    monkeypatch.delenv("SUPABASE_DATABASE_URL", raising=False)
    monkeypatch.delenv("DATABASE_URL", raising=False)
    from app import main
    from app.database import init_db

    monkeypatch.delenv("GOOGLE_CLIENT_ID", raising=False)
    main.AUTH_RATE_LIMIT = 0
    init_db()
    return TestClient(main.app)


def auth_headers(client: TestClient) -> dict[str, str]:
    header = base64.urlsafe_b64encode(json.dumps({"alg": "none"}).encode()).decode().rstrip("=")
    payload = base64.urlsafe_b64encode(
        json.dumps(
            {"sub": "excel-user", "email": "excel@example.com", "name": "excel", "exp": 4102444800}
        ).encode()
    ).decode().rstrip("=")
    login = client.post("/api/auth/google", json={"credential": f"{header}.{payload}."}).json()
    return {"Authorization": f"Bearer {login['session']['token']}"}


def workbook_payload(include_history: bool = False) -> dict[str, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Phase Tracker"
    sheet["A1"] = "FAT LOSS + HYBRID TRAINING - PHASE TRACKER"
    sheet["A2"] = "Goal: 88 kg → 72–73 kg | Priority: fat loss"
    headers = ["Phase", "Weight Range", "Calories*", "Protein", "Steps", "Running", "Main Goal"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=value)
    rows = [
        ["Phase 1", "88 → 84 kg", "~2050 kcal", "160–170 g", "~11k/day", "2 km Mon–Sat + 5 km Sun", "Build routine + strength"],
        ["Phase 2", "84 → 80 kg", "~2000 kcal*", "160–170 g", "~11k/day", "Same", "Continue fat loss + strength"],
        ["Phase 3", "80 → 76 kg", "~1900–2000*", "160–170 g", "~11k/day", "Same / adjust if needed", "Muscle preservation + recovery"],
        ["Phase 4", "76 → 72–73 kg", "~1850–1950*", "160–170 g", "~11k/day", "Same / controlled progression", "Final controlled cut"],
        ["Phase 5", "72–73 kg", "Maintenance", "150–165 g", "~11k/day", "Build toward 10K → HM → marathon", "Maintain + hybrid development"],
    ]
    for row_index, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)

    daily_headers = [
        "Day", "Date", "Weight", "Calories ✓", "Protein ✓", "Steps ✓", "Run ✓", "Gym ✓",
        "Breakfast ✓", "Lunch ✓", "Post-workout ✓", "Dinner ✓", "Sleep 8h ✓", "Notes",
    ]
    for column, value in enumerate(daily_headers, start=1):
        sheet.cell(row=13, column=column, value=value)
    row = 14
    for week in range(1, 9):
        sheet.cell(row=row, column=1, value=f"Week {week}")
        row += 1
        for day_name in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"):
            sheet.cell(row=row, column=1, value=day_name)
            for column in range(4, 14):
                sheet.cell(row=row, column=column, value="☐")
            row += 1
        row += 1
    if include_history:
        sheet.cell(row=15, column=2, value="2026-08-03")
        sheet.cell(row=15, column=3, value=88.0)
        sheet.cell(row=15, column=7, value="☐")
        sheet.cell(row=16, column=2, value="2026-08-04")
        sheet.cell(row=16, column=7, value="☑")
        sheet.cell(row=16, column=8, value="No")
        sheet.cell(row=16, column=13, value="✓")

    sheet["A87"] = "WEEKLY REVIEW"
    sheet["A99"] = "PHASE / ADJUSTMENT RULES"
    sheet["A100"] = "• Target loss: ~0.5–0.8 kg/week."
    sheet["A101"] = "• Move phases mainly by weight + waist + performance, not by calendar."
    buffer = BytesIO()
    workbook.save(buffer)
    return {
        "filename": "plan.xlsx",
        "file_base64": base64.b64encode(buffer.getvalue()).decode(),
        "start_date": "2026-08-03",
    }


def test_excel_plan_preview_understands_ranges_and_cycle(tmp_path, monkeypatch) -> None:
    client = make_client(tmp_path, monkeypatch)

    response = client.post(
        "/api/plan/import/excel/preview",
        json=workbook_payload(),
        headers=auth_headers(client),
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tracking_weeks"] == 8
    assert data["cycle_end_date"] == "2026-09-27"
    assert data["profile"] == {
        "start_weight_kg": 88.0,
        "goal_weight_min_kg": 72.0,
        "goal_weight_max_kg": 73.0,
    }
    assert data["phases"][2]["calorie_min"] == 1900
    assert data["phases"][2]["calorie_target"] == 1950
    assert data["phases"][3]["end_weight_kg"] == 73.0
    assert data["phases"][4]["calorie_target"] is None
    assert data["phases"][4]["weekday_run_km"] == 2.0


def test_excel_plan_apply_is_idempotent_and_effective_dated(tmp_path, monkeypatch) -> None:
    client = make_client(tmp_path, monkeypatch)
    headers = auth_headers(client)
    payload = workbook_payload()

    first = client.post("/api/plan/import/excel", json=payload, headers=headers)
    second = client.post("/api/plan/import/excel", json=payload, headers=headers)

    assert first.status_code == 200
    assert first.json()["applied"] is True
    assert second.json()["already_applied"] is True

    from app.database import connect

    with connect() as conn:
        phase_five = conn.execute("SELECT id FROM phases WHERE name = 'Phase 5'").fetchone()
        conn.execute("UPDATE user_profile SET current_phase_id = ? WHERE id = 1", (phase_five["id"],))
        conn.commit()

    before = client.get("/api/plan/timeline?date=2026-08-02", headers=headers).json()
    after = client.get("/api/plan/timeline?date=2026-08-03", headers=headers).json()
    assert before["current_phase"]["workout_days_per_week"] == 4
    assert after["current_phase"]["workout_days_per_week"] == 6
    assert after["cycle"]["current_week"] == 1


def test_excel_history_keeps_empty_checkboxes_unknown(tmp_path, monkeypatch) -> None:
    client = make_client(tmp_path, monkeypatch)
    headers = auth_headers(client)

    result = client.post(
        "/api/plan/import/excel",
        json=workbook_payload(include_history=True),
        headers=headers,
    ).json()

    assert result["daily_logs_imported"] == 2
    first = client.get("/api/day/2026-08-03", headers=headers).json()
    second = client.get("/api/day/2026-08-04", headers=headers).json()
    assert first["weight_kg"] == 88.0
    assert first["run_completed"] is None
    assert second["run_completed"] is True
    assert second["gym_completed"] is False
    assert second["sleep_hours"] == 8.0


def test_invalid_excel_payload_returns_clear_400(tmp_path, monkeypatch) -> None:
    client = make_client(tmp_path, monkeypatch)

    response = client.post(
        "/api/plan/import/excel/preview",
        json={"filename": "bad.xlsx", "file_base64": "not-base64", "start_date": "2026-08-03"},
        headers=auth_headers(client),
    )

    assert response.status_code == 400
    assert "base64" in response.json()["detail"]
